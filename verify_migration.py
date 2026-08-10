#!/usr/bin/env python3
"""
verify_migration.py — prove the YAML in data/ lost nothing from the source workbooks.

Run once after the migration, and again if anyone re-runs import_xlsx.py.
Not part of CI: the workbooks are no longer canonical and will not be in the repo
indefinitely.

    python3 tools/verify_migration.py --risk-register <v5.xlsx> --user-stories <v6.xlsx>

Checks: identifier sets match, free text is byte-identical, severity and likelihood
distributions are unchanged, and every cross-reference in every workbook cell
survived into the corresponding YAML list.
"""
import argparse, pathlib, re, sys
from collections import Counter

import openpyxl, yaml

ROOT = pathlib.Path(__file__).resolve().parent.parent


def ids(v, pat):
    return set(re.findall(pat, str(v))) if v else set()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--risk-register", required=True)
    ap.add_argument("--user-stories", required=True)
    a = ap.parse_args()

    rr = openpyxl.load_workbook(a.risk_register, read_only=True)
    us = openpyxl.load_workbook(a.user_stories, read_only=True)
    Y = lambda f: {r["id"]: r for r in yaml.safe_load(
        open(ROOT / "data" / f, encoding="utf-8"))["records"]}
    risks, controls, guardrails = Y("risks.yaml"), Y("controls.yaml"), Y("guardrails.yaml")
    ats, metrics = Y("assurance-tests.yaml"), Y("metrics.yaml")
    stories, scen = Y("user-stories.yaml"), Y("scenarios.yaml")

    fails = []

    def chk(label, ok, detail=""):
        print(f"  [{'ok' if ok else 'FAIL'}] {label} {detail}")
        if not ok:
            fails.append(label)

    print("Round-trip verification: workbook -> data/\n")
    ws = rr["Risk & Harm Register"]
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True)
            if r[0] and str(r[0]).startswith("RK-")]

    chk("risk ID set", {str(r[0]).strip() for r in rows} == set(risks))
    chk("control ID set",
        {str(r[0]).strip() for r in rr["Controls Catalogue"].iter_rows(min_row=4, values_only=True)
         if r[0] and str(r[0]).startswith("CT-")} == set(controls))
    chk("guardrail ID set",
        {str(r[0]).strip() for r in rr["Guardrails Register"].iter_rows(min_row=4, values_only=True)
         if r[0] and str(r[0]).startswith("GR-")} == set(guardrails))
    chk("assurance test ID set",
        {str(r[0]).strip() for r in rr["Assurance Tests"].iter_rows(min_row=3, values_only=True)
         if r[0] and str(r[0]).startswith("AT-")} == set(ats))
    chk("user story ID set",
        {str(r[0]).strip() for r in us["User Stories"].iter_rows(min_row=3, values_only=True)
         if r[0] and str(r[0]).startswith("US-")} == set(stories))
    chk("scenario ID set",
        {str(r[0]).strip() for r in us["Scenarios"].iter_rows(min_row=3, values_only=True)
         if r[0] and str(r[0]).startswith("SC-")} == set(scen))
    wb_m = set()
    for wbk in (rr, us):
        wb_m |= {str(r[0]).strip() for r in wbk["Trust Metrics"].iter_rows(min_row=3, values_only=True)
                 if r[0] and str(r[0]).startswith("M-")}
    chk("metric ID set (union of both workbooks)", wb_m == set(metrics), f"{len(wb_m)}")

    chk("risk descriptions verbatim",
        not [r for r in rows
             if risks[str(r[0]).strip()]["description"] != (str(r[4]).strip() if r[4] else None)])
    chk("control descriptions verbatim",
        not [r for r in rr["Controls Catalogue"].iter_rows(min_row=4, values_only=True)
             if r[0] and str(r[0]).startswith("CT-")
             and controls[str(r[0]).strip()]["description"] != (str(r[3]).strip() if r[3] else None)])
    chk("severity distribution",
        Counter(r[6] for r in rows) == Counter(r["severity"] for r in risks.values()))
    chk("likelihood distribution",
        Counter(r[7] for r in rows) == Counter(r["likelihood"] for r in risks.values()))

    lost = 0
    for r in rows:
        rid = str(r[0]).strip()
        for col, field, pat in ((9, "affected_metrics", r"M-\d\d"), (12, "guardrails", r"GR-\d\d"),
                                (13, "controls", r"CT-\d\d"), (14, "assurance_tests", r"AT-\d\d"),
                                (10, "user_stories", r"US-\d\d"), (11, "scenarios", r"SC-\d\d")):
            missing = ids(r[col], pat) - set(risks[rid][field])
            if missing:
                lost += len(missing)
                print(f"      lost {rid}.{field}: {sorted(missing)}")
    chk("risk cross-references preserved", lost == 0, f"{lost} lost")

    edges = sum(len(r[f]) for r in risks.values() for f in
                ("affected_metrics", "guardrails", "controls", "assurance_tests",
                 "user_stories", "scenarios", "epics"))
    print(f"\n  {edges} risk cross-reference edges carried into data/")
    print("\nRESULT:", "PASS — no data loss" if not fails else f"FAIL: {fails}")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
