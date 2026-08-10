#!/usr/bin/env python3
"""
import_xlsx.py — one-way migration from the RAHP Excel workbooks to canonical YAML.

This script exists for reproducibility and for the v0.3 migration only. Once the
YAML files under data/ are canonical, the workbooks become a *derived* view and
this script should not be run again. See CONTRIBUTING.md.

Usage:
    python3 tools/import_xlsx.py --risk-register <path.xlsx> \
                                 --user-stories <path.xlsx> \
                                 --out data/

Requires: openpyxl, PyYAML
"""
from __future__ import annotations

import argparse
import datetime
import pathlib
import re
import sys

try:
    import openpyxl
    import yaml
except ImportError:  # pragma: no cover
    sys.exit("Requires openpyxl and PyYAML: pip install -r requirements.txt")

ID_PATTERNS = {
    "RK": r"RK-[A-Z]{1,2}\d{2}",
    "CT": r"CT-\d{2}",
    "GR": r"GR-\d{2}",
    "AT": r"AT-\d{2}",
    "M": r"M-\d{2}",
    "US": r"US-\d{2}",
    "SC": r"SC-\d{2}",
    "EPIC": r"EPIC-\d{1,2}",
    "PERSONA": r"\b(?:D[1-6]|M[12]|B[123]|EC[1-4][ab]?)\b",
}

TODAY = datetime.date.today().isoformat()


def cell(v):
    """Normalise a cell to a stripped string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def ids(v, kind):
    """Extract all identifiers of `kind` from a cell, de-duplicated, order preserved."""
    if v is None:
        return []
    found = re.findall(ID_PATTERNS[kind], str(v))
    seen, out = set(), []
    for f in found:
        if f not in seen:
            seen.add(f)
            out.append(f)
    return out


def rows(ws, header_row):
    return list(ws.iter_rows(min_row=header_row + 1, values_only=True))


def slug_priority(v):
    if not v:
        return "unassigned"
    s = str(v).lower()
    if "must" in s:
        return "must_address"
    if "should" in s:
        return "should_address"
    if "monitor" in s:
        return "monitor"
    return "unassigned"


def provenance(source_file, note=None):
    p = {"source": f"workbook:{source_file}", "imported": TODAY}
    if note:
        p["note"] = note
    return p


# --------------------------------------------------------------------------
# Record builders
# --------------------------------------------------------------------------

def build_risks(wb, src):
    ws = wb["Risk & Harm Register"]
    out = []
    for r in rows(ws, 3):
        rid = cell(r[0])
        if not rid or not rid.startswith("RK-"):
            continue
        out.append({
            "id": rid,
            "name": cell(r[1]),
            "category": cell(r[2]),
            "lifecycle_phase": cell(r[3]),
            "description": cell(r[4]),
            "harm_types": [h.strip() for h in re.split(r"[;/]", cell(r[5]) or "") if h.strip()],
            "severity": cell(r[6]),
            "likelihood": cell(r[7]),
            "harm_description": cell(r[8]),
            "affected_metrics": ids(r[9], "M"),
            "user_stories": ids(r[10], "US"),
            "scenarios": ids(r[11], "SC"),
            "guardrails": ids(r[12], "GR"),
            "controls": ids(r[13], "CT"),
            "assurance_tests": ids(r[14], "AT"),
            "epics": ids(r[15], "EPIC"),
            "standards_priority": slug_priority(r[16]),
            "provenance": provenance(src),
        })
    return out


def build_controls(wb, src):
    ws = wb["Controls Catalogue"]
    out = []
    for r in rows(ws, 3):
        cid = cell(r[0])
        if not cid or not cid.startswith("CT-"):
            continue
        out.append({
            "id": cid,
            "name": cell(r[1]),
            "type": cell(r[2]),
            "description": cell(r[3]),
            "guardrails": ids(r[4], "GR"),
            "linked_risks": ids(r[5], "RK"),
            "standards_relevance": cell(r[6]),
            # v0.3 additions — deliberately unassigned until the task force rules on them.
            "standards_status": "unassigned",
            "normative_language": None,
            "rationale": cell(r[7]),
            "provenance": provenance(src),
        })
    return out


def build_guardrails(wb, src):
    ws = wb["Guardrails Register"]
    out = []
    for r in rows(ws, 3):
        gid = cell(r[0])
        if not gid or not gid.startswith("GR-"):
            continue
        out.append({
            "id": gid,
            "name": cell(r[1]),
            "category": cell(r[2]),
            "requirement": cell(r[3]),
            "applies_to_phase": cell(r[4]),
            "owner": cell(r[5]),
            "risks_addressed": ids(r[6], "RK"),
            "controls": ids(r[7], "CT"),
            "assurance_tests": ids(r[8], "AT"),
            "standards_status": "unassigned",
            "normative_language": None,
            "provenance": provenance(src),
        })
    return out


def build_assurance_tests(wb, src):
    ws = wb["Assurance Tests"]
    out = []
    for r in rows(ws, 2):
        aid = cell(r[0])
        if not aid or not aid.startswith("AT-"):
            continue
        out.append({
            "id": aid,
            "guardrail": (ids(r[1], "GR") or [None])[0],
            "pass_criterion": cell(r[2]),
            "verification_method": cell(r[3]),
            "automated": bool(re.search(r"automat", str(r[3] or ""), re.I)),
            "test_type": cell(r[4]),
            "risks_covered": ids(r[5], "RK"),
            "notes": cell(r[6]),
            "provenance": provenance(src),
        })
    return out


def build_metrics(rr, us, src_rr, src_us):
    """
    Metrics are the shared identifier space. v5/v6 disagree: the Risk Register
    defines M-01..M-18 + M-31..M-37, the User Stories Framework defines M-01..M-37.
    We take the union, preferring the User Stories Framework definition where both
    exist, and record which workbook each definition came from.
    """
    defs = {}
    for wb, src, sheet in ((rr, src_rr, "Trust Metrics"), (us, src_us, "Trust Metrics")):
        for r in rows(wb[sheet], 2):
            mid = cell(r[0])
            if not mid or not mid.startswith("M-"):
                continue
            defs[mid] = {
                "id": mid,
                "name": cell(r[1]),
                "category": cell(r[2]),
                "description": cell(r[3]),
                "personas_want_high": ids(r[4], "PERSONA"),
                "personas_want_suppressed": ids(r[5], "PERSONA"),
                "provenance": provenance(src),
            }

    # risks measured — from the Risk Register Metric Detail sheet
    for r in rows(rr["Metric Detail"], 2):
        mid = cell(r[0])
        if mid in defs:
            defs[mid]["risks_measured"] = ids(r[3], "RK")

    # US / SC / EPIC triggers — from the User Stories Metric Cross-Reference sheet
    for r in rows(us["Metric Cross-Reference"], 2):
        mid = cell(r[0])
        if mid in defs:
            defs[mid]["user_stories"] = ids(r[3], "US")
            defs[mid]["scenarios"] = ids(r[4], "SC")
            defs[mid]["epics"] = ids(r[5], "EPIC")

    out = []
    for mid in sorted(defs, key=lambda x: int(x.split("-")[1])):
        d = defs[mid]
        d.setdefault("risks_measured", [])
        d.setdefault("user_stories", [])
        d.setdefault("scenarios", [])
        d.setdefault("epics", [])
        # v0.4 hook: runtime monitoring definition, unpopulated at v0.3.
        d["monitoring"] = None
        out.append(d)
    return out


def build_user_stories(wb, src):
    ws = wb["User Stories"]
    out = []
    for r in rows(ws, 2):
        uid = cell(r[0])
        if not uid or not uid.startswith("US-"):
            continue
        persona_cell = cell(r[1]) or ""
        out.append({
            "id": uid,
            "persona": (ids(persona_cell, "PERSONA") or [None])[0],
            "persona_label": persona_cell or None,
            "function": cell(r[2]),
            "goal": cell(r[3]),
            "lifecycle_phase": cell(r[4]),
            "scenarios": ids(r[5], "SC"),
            "epics": ids(r[6], "EPIC"),
            "risk_categories": [c.strip() for c in re.split(r",", cell(r[7]) or "") if c.strip()],
            "metrics": ids(r[8], "M"),
            "provenance": provenance(src),
        })
    return out


def build_scenarios(wb, src):
    ws = wb["Scenarios"]
    out = []
    for r in rows(ws, 2):
        sid = cell(r[0])
        if not sid or not sid.startswith("SC-"):
            continue
        out.append({
            "id": sid,
            "name": cell(r[1]),
            "description": cell(r[2]),
            "lifecycle_phase": cell(r[3]),
            "legitimate_personas": ids(r[4], "PERSONA"),
            "adversarial_personas": ids(r[5], "PERSONA"),
            "governance_decisions": cell(r[6]),
            "user_stories": ids(r[7], "US"),
            "metrics": ids(r[8], "M"),
            "provenance": provenance(src),
        })
    return out


def build_epics(wb, src):
    ws = wb["EPICs"]
    out = []
    for r in rows(ws, 2):
        eid = cell(r[0])
        if not eid or not eid.startswith("EPIC-"):
            continue
        out.append({
            "id": eid,
            "name": cell(r[1]),
            "layer": cell(r[2]),
            "description": cell(r[3]),
            "personas": ids(r[4], "PERSONA"),
            "user_stories": ids(r[5], "US"),
            "metrics": ids(r[6], "M"),
            "provenance": provenance(src),
        })
    return out


def build_persona_index(wb, src):
    """Cross-reference rows only. Narrative persona content lives in data/personas.yaml
    and is maintained by hand — it is not workbook-derived."""
    ws = wb["Persona Cross-Reference"]
    out = {}
    for r in rows(ws, 2):
        pid = cell(r[0])
        if not pid or pid == "Persona ID":
            continue
        out[pid] = {
            "name": cell(r[1]),
            "type": cell(r[2]),
            "user_stories": ids(r[3], "US"),
            "scenarios": ids(r[4], "SC"),
            "epics": ids(r[5], "EPIC"),
            "metrics": ids(r[6], "M"),
        }
    return out


def dump(path, kind, records, extra_header=None):
    payload = {
        "record_type": kind,
        "generated_by": "tools/import_xlsx.py",
        "records": records,
    }
    if extra_header:
        payload.update(extra_header)
    with open(path, "w", encoding="utf-8") as fh:
        yaml.safe_dump(payload, fh, sort_keys=False, allow_unicode=True, width=100)
    print(f"  {path.name}: {len(records)} records")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--risk-register", required=True)
    ap.add_argument("--user-stories", required=True)
    ap.add_argument("--out", default="data")
    a = ap.parse_args()

    out = pathlib.Path(a.out)
    out.mkdir(parents=True, exist_ok=True)

    rr_src = pathlib.Path(a.risk_register).name
    us_src = pathlib.Path(a.user_stories).name
    rr = openpyxl.load_workbook(a.risk_register, read_only=True, data_only=True)
    us = openpyxl.load_workbook(a.user_stories, read_only=True, data_only=True)

    print("Importing:")
    dump(out / "risks.yaml", "risk", build_risks(rr, rr_src))
    dump(out / "controls.yaml", "control", build_controls(rr, rr_src))
    dump(out / "guardrails.yaml", "guardrail", build_guardrails(rr, rr_src))
    dump(out / "assurance-tests.yaml", "assurance_test", build_assurance_tests(rr, rr_src))
    dump(out / "metrics.yaml", "metric", build_metrics(rr, us, rr_src, us_src))
    dump(out / "user-stories.yaml", "user_story", build_user_stories(us, us_src))
    dump(out / "scenarios.yaml", "scenario", build_scenarios(us, us_src))
    dump(out / "epics.yaml", "epic", build_epics(us, us_src))

    # Persona cross-reference is merged into the hand-maintained personas.yaml
    # by tools/merge_persona_xrefs.py rather than overwriting it here.
    idx = build_persona_index(us, us_src)
    with open(out / "_persona-xref.yaml", "w", encoding="utf-8") as fh:
        yaml.safe_dump({"record_type": "persona_xref", "records": idx},
                       fh, sort_keys=False, allow_unicode=True, width=100)
    print(f"  _persona-xref.yaml: {len(idx)} records")


if __name__ == "__main__":
    main()
